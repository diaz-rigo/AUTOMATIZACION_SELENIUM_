import psycopg2
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# Configuración de rango
RANGO_INICIO = 1
RANGO_FIN = 5

# Conexión a la base de datos PostgreSQL
try:
    conn = psycopg2.connect(
        dbname="erpp",
        user="angel",
        password="ErppHgo&2024",
        host="100.66.168.122",
        port="5432"
    )
    cursor = conn.cursor()
except Exception as e:
    print(f"Error conectando a la base de datos: {e}")
    exit()

# Cargar el archivo Excel
file_path = r'/Users/arturo/Documents/EXTINTOS_SCRIP_ UPDATE/Hazaelfolios.xlsx'

# Leer la hoja 'Masivos' del archivo Excel
df = pd.read_excel(file_path, sheet_name='Masivos')

# Verificar si las columnas 'FOLIOREAL' y 'CARGA' existen
if 'FOLIOREAL' in df.columns and 'CARGA' in df.columns:
    # Filtrar las filas por el rango de 'CARGA'
    df_filtrado = df[(df['CARGA'] >= RANGO_INICIO) & (df['CARGA'] <= RANGO_FIN)]

    # Lista de columnas a tomar de Excel y verificar en la base de datos
    columnas_excel = [
        'manzana', 'no_lote', 'nombre_frac_o_condo', 'num_ext', 'num_int',
        'superficie', 'superficie_m2', 'vialidad', 'municipio_id', 'tipo_asent_id', 'tipo_inmueble_id'
    ]

    # Crear listas para almacenar los resultados
    datos_excel = []
    datos_bd = []

    # Recorrer y procesar los folios filtrados
    if not df_filtrado.empty:
        for index, row in df_filtrado.iterrows():
            # Imprimir el número de carga actual
            print(f"Procesando carga: {row['CARGA']} (fila {index + 2})")

            try:
                # Convertir el folio a entero (elimina decimales como 672539.0 -> 672539)
                folio = int(float(str(row['FOLIOREAL']).strip()))
            except ValueError:
                print(f"Error al convertir folio en la fila {index + 2}")
                continue

            # Consultar los datos del folio en la base de datos
            cursor.execute("""
                SELECT *, COUNT(*) OVER () AS duplicados 
                FROM predio 
                WHERE no_folio = %s
            """, (folio,))
            predio_data = cursor.fetchone()

            # Si se encuentran datos en la BD
            if predio_data:
                duplicados = predio_data[-1]  # La columna duplicados es la última

                # Extraer los datos de las columnas específicas del Excel
                datos_folio_excel = [row[columna] for columna in columnas_excel]

                # Agregar datos encontrados de la BD y Excel a sus respectivas listas
                datos_bd.append([folio, row['CARGA'], duplicados] + list(predio_data[:-1]))
                datos_excel.append([folio, row['CARGA']] + datos_folio_excel)

            # Si no se encuentra en la base de datos
            else:
                datos_excel.append([folio, row['CARGA']] + [row[col] for col in columnas_excel] + ["No encontrado en BD"])
                datos_bd.append([folio, row['CARGA'], "No encontrado"])

        # Crear un nuevo archivo Excel con dos hojas
        wb = Workbook()
        ws_excel = wb.active
        ws_excel.title = "Datos Excel"
        ws_bd = wb.create_sheet(title="Datos BD")

        # Definir columnas para ambas hojas
        columnas_bd = [desc[0] for desc in cursor.description[:-1]] if cursor.description else []
        columnas_resultado_bd = ['folio', 'carga', 'duplicados'] + columnas_bd
        columnas_resultado_excel = ['folio', 'carga'] + columnas_excel + ['estado_en_bd']

        # Escribir los encabezados en ambas hojas
        ws_excel.append(columnas_resultado_excel)
        ws_bd.append(columnas_resultado_bd)

        # Escribir los datos en cada hoja
        for fila_excel, fila_bd in zip(datos_excel, datos_bd):
            ws_excel.append(fila_excel)
            ws_bd.append(fila_bd)

        # Resaltar duplicados en la hoja de BD (color amarillo)
        fill_duplicado = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in ws_bd.iter_rows(min_row=2, max_row=ws_bd.max_row, min_col=3, max_col=3):
            for cell in row:
                if isinstance(cell.value, int) and cell.value > 1:
                    cell.fill = fill_duplicado  # Resaltar la celda con duplicados

        # Guardar el nuevo archivo Excel
        new_file_path = r'/Users/arturo/Documents/EXTINTOS_SCRIP_ UPDATE/Nuevo_Excel_Consultas_Folios.xlsx'
        wb.save(new_file_path)

        print(f"Datos procesados y guardados en: {new_file_path}")

    else:
        print(f"No se encontraron filas en el rango de CARGA {RANGO_INICIO} - {RANGO_FIN}.")
else:
    print("Las columnas 'FOLIOREAL' o 'CARGA' no existen en la hoja 'Masivos'.")

# Cerrar la conexión a la base de datos
cursor.close()
conn.close()
