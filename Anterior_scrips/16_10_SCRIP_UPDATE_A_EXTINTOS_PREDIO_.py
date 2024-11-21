import psycopg2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

RANGO_INICIO = 1
RANGO_FIN = 39702
# teermino en folio  1605213
# Conexión a la base de datos PostgreSQL
conn = psycopg2.connect(
    dbname="erpp",
    user="angel",
    password="ErppHgo&2024",
    host="100.66.168.122",
    port="5432"
)
cursor = conn.cursor()

# Cargar el archivo Excel
file_path = r'/Users/arturo/Documents/EXTINTOS_SCRIP_ UPDATE/Libro 2_Extinguir.xlsx'

# Leer la hoja 'FINAL' del archivo Excel
df = pd.read_excel(file_path, sheet_name='FINAL')

# Verificar si las columnas 'FOLIOREAL' y 'CARGA' existen
if 'FOLIOREAL' in df.columns and 'CARGA' in df.columns:
    # Filtrar las filas por el rango de 'CARGA'
    df_filtrado = df[(df['CARGA'] >= RANGO_INICIO) & (df['CARGA'] <= RANGO_FIN)]

    # Cargar el archivo Excel para modificarlo
    wb = load_workbook(file_path)
    ws = wb['FINAL']

    # Relleno para las celdas que se actualizarán
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Agregar una nueva columna para el estado anterior y otra para el nuevo estado si no existen
    if 'STATUS_ANTERIOR' not in df.columns:
        df['STATUS_ANTERIOR'] = None  # Inicializar la columna como None
    if 'STATUS_NUEVO' not in df.columns:
        df['STATUS_NUEVO'] = None  # Inicializar la columna para el nuevo status

    # Recorrer e imprimir los valores filtrados de la columna 'FOLIOREAL'
    if not df_filtrado.empty:
        folios_actualizados = []  # Almacenar folios que han sido actualizados
        for index, folio in enumerate(df_filtrado['FOLIOREAL']):
            folio = str(folio).strip()  # Eliminar espacios al inicio y final del folio

            # Obtener el estado actual del folio
            cursor.execute(f"SELECT status_acto_id FROM predio WHERE no_folio = '{folio}';")
            current_status = cursor.fetchone()

            if current_status:  # Verificar si se encontró un estado
                # Guardar el estado anterior en el DataFrame
                df.loc[df['FOLIOREAL'] == folio, 'STATUS_ANTERIOR'] = current_status[0]

                # Actualizar el estado en la base de datos
                update_query = f"UPDATE predio SET status_acto_id = 4 WHERE no_folio = '{folio}';"
                print(update_query)
                cursor.execute(update_query)
                folios_actualizados.append(folio)  # Agregar folio actualizado a la lista
                print("folio:::::::::::::::::::::::::::::::::::::",folio)
                # Guardar el nuevo estado en el DataFrame
                df.loc[df['FOLIOREAL'] == folio, 'STATUS_NUEVO'] = 4  # Asignar el nuevo status

        # Ejecutar un solo commit para todos los updates
        conn.commit()

        # Aplicar el color a las filas de los folios actualizados
        for folio in folios_actualizados:
            fila_excel = df.index[df['FOLIOREAL'] == folio].tolist()[0] + 2  # +2 por encabezados y el índice de Excel
            for col in range(1, len(df.columns) + 1):  # Recorre todas las columnas
                ws.cell(row=fila_excel, column=col).fill = highlight_fill

        # Guardar los cambios en el archivo Excel
        df.to_excel(file_path, sheet_name='FINAL', index=False)  # Guardar el DataFrame en el Excel
        wb.save(file_path)  # Asegurarse de guardar los cambios de estilo en el archivo Excel
        print("Cambios guardados en el archivo Excel.")
    else:
        print(f"No se encontraron filas en el rango de CARGA {RANGO_INICIO} - {RANGO_FIN}.")
else:
    print("Las columnas 'FOLIOREAL' o 'CARGA' no existen en la hoja 'FINAL'.")

# Cerrar la conexión
cursor.close()
conn.close()
