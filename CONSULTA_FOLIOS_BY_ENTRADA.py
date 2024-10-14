import psycopg2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Filtrar las filas por la columna ENTRADA y por rango de CARGA
ENTRADA_FILTRO = 52872
RANGO_INICIO = 1
RANGO_FIN = 3

# Conexión a la base de datos PostgreSQL
conn = psycopg2.connect(
    dbname="erpp",
    user="angel",
    password="ErppHgo&2024",
    host="100.66.168.122",
    port="5432"
)
cursor = conn.cursor()

# Cargar el archivo Excel
file_path = '/Users/hazael/Documents/SCRIP_ UPDATE/RIGOFOLIOSBUENO.xlsx'
df = pd.read_excel(file_path, sheet_name=str(ENTRADA_FILTRO))

# Función para consultar la entrada en la tabla `prelacion`
def consultar_prelacion(consecutivo, anio):
    consulta = """
        SELECT * FROM prelacion p 
        WHERE p.consecutivo = %s AND anio = %s;
    """
    cursor.execute(consulta, (consecutivo, anio))
    resultado = cursor.fetchall()
    print(f"Consulta prelacion: {consulta} con valores ({consecutivo}, {anio})")
    return resultado if resultado else None

# Función para consultar en la tabla `prelacion_ante` con parámetros adicionales
def prelacion_ante_consultar_documento(prelacion_id, documento, anio, libro, volumen):
    consulta = """
        SELECT * FROM prelacion_ante pa 
        WHERE prelacion_id = %s AND documento = %s AND anio = %s AND libro = %s AND volumen = %s;
    """
    cursor.execute(consulta, (prelacion_id, documento, anio, libro, volumen))
    resultados = cursor.fetchall()
    print(f"Consulta documento: {consulta} con valores ({prelacion_id}, {documento}, {anio}, {libro}, {volumen})")

    # Si hay más de un resultado, se considera duplicado
    if len(resultados) > 1:
        return None, "Duplicado", len(resultados)

    # Si no hay resultados, retorna No encontrado
    if not resultados:
        return None, "No encontrado", 0

    # Si se encontró exactamente un resultado
    return resultados[0], "Encontrado", 1

# Función para consultar en la tabla `predio` y devolver 'no_folio'
def consultar_predio(predio_id):
    consulta = """
        SELECT id, no_folio FROM predio WHERE id = %s;
    """
    cursor.execute(consulta, (predio_id,))
    resultado = cursor.fetchone()
    print(f"Consulta predio: {consulta} con valor predio_id = {predio_id}")
    return resultado if resultado else None

# Función para actualizar el campo 'no_folio' en la tabla `predio`
def actualizar_folio(predio_id, new_folio_excel):
    consulta = """
        UPDATE predio 
        SET no_folio = %s
        WHERE id = %s;
    """
    cursor.execute(consulta, (new_folio_excel, predio_id))
    conn.commit()
    print(f"Update predio: {consulta} con valores (new_folio_excel = {new_folio_excel}, predio_id = {predio_id})")

# Filtrar el DataFrame según los criterios usando 'CARGA' en lugar de 'CONSECUTIVO'
df_filtrado = df[(df['CARGA'] >= RANGO_INICIO) & (df['CARGA'] <= RANGO_FIN)]

# Imprimir las columnas del Excel
print("Columnas del archivo Excel:", df.columns)

# Crear una nueva columna para 'no_folio', 'indicador_duplicado' y 'cantidad_duplicados' en el DataFrame
df['no_folio'] = None
df['indicador_duplicado'] = None
df['cantidad_duplicados'] = None
df['predio_id'] = None  # Nueva columna para guardar el predio_id
df['nuevo_folio'] = None  # Nueva columna para el nuevo folio

for index, row in df_filtrado.iterrows():
    indicador_duplicado = "Sin duplicado"
    cantidad_duplicados = 0
    prelacion = consultar_prelacion(ENTRADA_FILTRO, '2024')
    print("prelacion-----", prelacion)

    if prelacion:
        prelacion_id = prelacion[0][0]
        documento = str(row['INSCRIPCION'])
        libro = str(row['TOMO'])
        volumen = str(row['VOLUMEN'])
        anio = row['AÑO']
        antecedente, indicador_duplicado, cantidad_duplicados = prelacion_ante_consultar_documento(
            prelacion_id, documento, anio, libro, volumen)
        print("---antecedente", antecedente)
        print(indicador_duplicado)

        if antecedente:
            predio_id = antecedente[13]

            # Consultar predio
            predio = consultar_predio(predio_id)
            print("predio..", predio)
            if predio:
                no_folio = predio[1]
                print("no_folio***", no_folio)

                # Actualizar las columnas en el DataFrame
                df.at[index, 'no_folio'] = no_folio
                df.at[index, 'predio_id'] = predio_id

                # Verificar si 'FOLIO REAL' está presente en la fila
                new_folio = row.get('FOLIO REAL', None)

                if new_folio:
                    df.at[index, 'nuevo_folio'] = new_folio
                    # Actualizar el folio en la base de datos
                    # **************ACTUALIZAR FOLIO
                    # actualizar_folio(predio_id, new_folio)
                else:
                    print(f"Folio REAL no disponible para predio_id: {predio_id} en la fila {index}")
            else:
                df.at[index, 'no_folio'] = "Predio no encontrado"
                df.at[index, 'predio_id'] = "N/A"
        else:
            df.at[index, 'no_folio'] = "Antecedente no encontrado"
            df.at[index, 'predio_id'] = "N/A"
    else:
        df.at[index, 'no_folio'] = "Prelacion no encontrada"
        df.at[index, 'predio_id'] = "N/A"

    # Actualizar los valores en el DataFrame
    df.at[index, 'indicador_duplicado'] = indicador_duplicado
    df.at[index, 'cantidad_duplicados'] = cantidad_duplicados
    print("-----------------------------------------------------------------------------------")

# Guardar el DataFrame de vuelta en el mismo archivo Excel existente
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
    # Borrar la hoja existente antes de volver a escribir
    if str(ENTRADA_FILTRO) in writer.book.sheetnames:
        del writer.book[str(ENTRADA_FILTRO)]
    df.to_excel(writer, sheet_name=str(ENTRADA_FILTRO), index=False)

# Cargar el archivo con `openpyxl` para aplicar formato
wb = load_workbook(file_path)
ws = wb[str(ENTRADA_FILTRO)]

# Crear un relleno de color para los valores encontrados (color azul claro)
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
# Crear un relleno de color para los valores no encontrados (color rojo claro)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Buscar la columna de 'no_folio'
no_folio_column = None
for idx, column in enumerate(ws.iter_cols(1, ws.max_column)):
    if column[0].value == 'no_folio':
        no_folio_column = idx + 1  # Columna comienza en 1
        break

# Aplicar color a las celdas en la columna "no_folio"
if no_folio_column:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=no_folio_column, max_col=no_folio_column):
        for cell in row:
            if cell.value in ["Predio no encontrado", "Antecedente no encontrado", "Prelacion no encontrada"]:
                cell.fill = red_fill
            # Valores válidos (suponiendo que sean números o textos)
            elif cell.value not in [None, ""]:
                cell.fill = blue_fill

# Guardar el archivo Excel con formato
wb.save(file_path)
print("Archivo Excel actualizado y guardado con color en la columna 'no_folio'.")
